"""Statistics API routes."""

from datetime import datetime, date, timedelta
from typing import Optional
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, text, literal_column

from app.database import get_db
from app.models.waybill import LogisticsItem, LogisticsStatus
from app.models.region import Region

router = APIRouter(prefix="/stats", tags=["통계"])


@router.get("/regions", response_model=dict)
async def get_region_stats(
    date: Optional[str] = Query(None, description="조회 날짜 (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """STAT-001: 구역별 현황 조회."""
    query = db.query(
        LogisticsItem.destination,
        func.count(LogisticsItem.tracking_number).label("total"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.READY, 1), else_=0)).label("ready"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.MOVING, 1), else_=0)).label("moving"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.COMPLETED, 1), else_=0)).label("completed"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.ERROR, 1), else_=0)).label("error")
    )
    
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
            # Use range query for better index usage and reliability
            start_dt = datetime.combine(target_date, datetime.min.time())
            end_dt = datetime.combine(target_date, datetime.max.time())
            
            # Adjust query to filter by range
            query = query.filter(LogisticsItem.created_at >= start_dt, LogisticsItem.created_at <= end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail={
                "code": "INVALID_DATE",
                "message": "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)"
            })
    
    query = query.filter(LogisticsItem.destination.isnot(None))
    results = query.group_by(LogisticsItem.destination).all()
    
    stats = []
    for row in results:
        stats.append({
            "region_id": row.destination,
            "region_name": row.destination, # Use raw destination from DB directly
            "total": row.total,
            "ready": row.ready,
            "moving": row.moving,
            "completed": row.completed,
            "error": row.error
        })
    
    return {
        "success": True,
        "data": stats
    }


@router.get("/daily", response_model=dict)
async def get_daily_stats(
    start_date: Optional[str] = Query(None, description="시작 날짜 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="종료 날짜 (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """STAT-002: 일별 처리 통계 조회."""
    query = db.query(
        func.date(LogisticsItem.created_at).label("date"),
        func.count(LogisticsItem.tracking_number).label("total"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.COMPLETED, 1), else_=0)).label("completed"),
        func.sum(case((LogisticsItem.status == LogisticsStatus.ERROR, 1), else_=0)).label("error"),
        func.avg(
            case(
                (LogisticsItem.completed_at.isnot(None),
                 func.timestampdiff(literal_column("SECOND"), LogisticsItem.created_at, LogisticsItem.completed_at)),
                else_=None
            )
        ).label("avg_process_time")
    )
    
    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(func.date(LogisticsItem.created_at) >= start)
        except ValueError:
            raise HTTPException(status_code=400, detail={
                "code": "INVALID_DATE",
                "message": "시작 날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)"
            })
    
    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(func.date(LogisticsItem.created_at) <= end)
        except ValueError:
            raise HTTPException(status_code=400, detail={
                "code": "INVALID_DATE",
                "message": "종료 날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)"
            })
    
    results = query.group_by(func.date(LogisticsItem.created_at)).order_by(func.date(LogisticsItem.created_at).desc()).all()
    
    stats = []
    for row in results:
        stats.append({
            "date": str(row.date),
            "total": row.total,
            "completed": row.completed,
            "error": row.error,
            "avg_process_time_sec": int(row.avg_process_time) if row.avg_process_time else None
        })
    
    return {
        "success": True,
        "data": stats
    }


@router.get("/export")
async def export_stats(
    date: Optional[str] = Query(None, description="조회 날짜 (YYYY-MM-DD)"),
    start_date: Optional[str] = Query(None, description="시작 날짜 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="종료 날짜 (YYYY-MM-DD)"),
    db: Session = Depends(get_db)
):
    """STAT-004: 물류 데이터 엑셀 다운로드 (단일 날짜, 기간, 또는 전체)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail={
            "code": "EXPORT_ERROR",
            "message": "Excel 라이브러리가 설치되지 않았습니다"
        })
    
    query = db.query(LogisticsItem)
    
    # Date Filtering Logic
    try:
        if start_date and end_date:
            # Range filter
            s_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            e_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            # Include the entire end date (up to 23:59:59.999999 if using datetime comparison on created_at)
            # Since created_at is timestamp, we filter by date(created_at) or use range
            query = query.filter(func.date(LogisticsItem.created_at) >= s_date, func.date(LogisticsItem.created_at) <= e_date)
            filename_date = f"{start_date}_{end_date}"
            
        elif date:
            # Single date filter
            filter_date = datetime.strptime(date, "%Y-%m-%d").date()
            query = query.filter(func.date(LogisticsItem.created_at) == filter_date)
            filename_date = date
            
        else:
            # No date filter -> All data
            filename_date = datetime.now().strftime('%Y%m%d')
            
    except ValueError:
        raise HTTPException(status_code=400, detail={
            "code": "INVALID_DATE",
            "message": "날짜 형식이 올바르지 않습니다 (YYYY-MM-DD)"
        })
    
    items = query.order_by(LogisticsItem.created_at.desc()).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "물류 데이터"
    
    # Headers
    headers = ["운송장번호", "목적지", "상태", "생성일시", "완료일시", "처리시간(초)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    for row, item in enumerate(items, 2):
        process_time = None
        if item.completed_at and item.created_at:
            process_time = int((item.completed_at - item.created_at).total_seconds())
        
        ws.cell(row=row, column=1, value=item.tracking_number)
        ws.cell(row=row, column=2, value=item.destination)
        ws.cell(row=row, column=3, value=item.status.value)
        ws.cell(row=row, column=4, value=str(item.created_at) if item.created_at else None)
        ws.cell(row=row, column=5, value=str(item.completed_at) if item.completed_at else None)
        ws.cell(row=row, column=6, value=process_time)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"logistics_export_{filename_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
